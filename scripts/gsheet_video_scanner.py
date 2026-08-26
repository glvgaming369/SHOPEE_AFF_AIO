"""Scan a video folder and match it against a *_results.xlsx export."""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string

VIDEO_FILENAME_RE = re.compile(r"^(\d+)\.mp4$", re.IGNORECASE)

_SRC_COL_SP_ID = column_index_from_string("A") - 1
_SRC_COL_PRODUCT_NAME = column_index_from_string("B") - 1
_SRC_COL_SHOPEE_URL = column_index_from_string("C") - 1
_SRC_COL_MERGE_LINKS = column_index_from_string("P") - 1


@dataclass(frozen=True)
class ProductRow:
    sp_id: str
    product_name: str
    shopee_url: str
    merge_links: str


def scan_video_ids(folder: Path) -> set[str]:
    """Return SP IDs of files matching exactly `<digits>.mp4` in folder (non-recursive)."""
    ids: set[str] = set()
    for entry in folder.iterdir():
        if not entry.is_file():
            continue
        match = VIDEO_FILENAME_RE.match(entry.name)
        if match:
            ids.add(match.group(1))
    return ids


def find_results_file(folder: Path) -> Path | None:
    matches = sorted(folder.glob("*_results.xlsx"))
    if len(matches) > 1:
        names = ", ".join(m.name for m in matches)
        raise ValueError(
            f"Tìm thấy nhiều file *_results.xlsx trong thư mục ({names}). "
            "Vui lòng chỉ để lại đúng 1 file."
        )
    return matches[0] if matches else None


def _normalize_sp_id(value) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_product_rows(results_path: Path) -> list[ProductRow]:
    """Read data rows (skipping the header) from the first worksheet, top to bottom."""
    workbook = openpyxl.load_workbook(results_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        if worksheet.max_column <= _SRC_COL_MERGE_LINKS:
            raise ValueError(
                f"File {results_path.name} chỉ có {worksheet.max_column} cột, thiếu cột "
                "P (Merge Links) theo đúng định dạng mong đợi."
            )
        rows: list[ProductRow] = []
        for raw in worksheet.iter_rows(min_row=2, values_only=True):
            if raw is None or len(raw) <= _SRC_COL_MERGE_LINKS:
                continue
            sp_id = raw[_SRC_COL_SP_ID]
            if sp_id is None or str(sp_id).strip() == "":
                continue
            rows.append(
                ProductRow(
                    sp_id=_normalize_sp_id(sp_id),
                    product_name=str(raw[_SRC_COL_PRODUCT_NAME] or "").strip(),
                    shopee_url=str(raw[_SRC_COL_SHOPEE_URL] or "").strip(),
                    merge_links=str(raw[_SRC_COL_MERGE_LINKS] or "").strip(),
                )
            )
        return rows
    finally:
        workbook.close()


def build_matched_pool(folder: str | Path) -> list[ProductRow]:
    """Videos on disk AND rows in *_results.xlsx, intersected, ordered as in the xlsx."""
    folder = Path(folder)
    results_path = find_results_file(folder)
    if results_path is None:
        raise FileNotFoundError(f"Không tìm thấy file *_results.xlsx trong {folder}")
    video_ids = scan_video_ids(folder)
    rows = read_product_rows(results_path)
    return [row for row in rows if row.sp_id in video_ids]
