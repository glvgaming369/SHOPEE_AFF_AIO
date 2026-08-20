"""Tra 'cat_name' (ten category hien thi) tu cat_id + market, doc tu
artifacts/cat-db/*.xlsx (sheet 'Shopee_Data': STT, cat_name, cat_url, cat_id - moi file la
danh sach category cap 1 cua 1 thi truong). Dung cho:
    - import_roots_as_pending() (shopee_db.py): gan cat_name hien thi khi Shopee Product
      Link Collector day cat_id phat hien tu URL len.
    - try_assign_verified() (shopee_db.py): ke thua lai cat_name cua root khi 1 san pham
      tuong tu duoc gan vao group.

Cache trong RAM, CHI doc file 1 LAN (lazy, lan goi dau tien) - sua file cat-db/*.xlsx xong
phai KHOI DONG LAI affiliate_scrape_server.py moi nhan duoc thay doi (server chay
debug=False, khong tu dong reload gi ca - giong cach template index.html cung can restart,
xem CLAUDE.md/session note ve van de nay)."""
import os

import openpyxl

CAT_DB_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "artifacts", "cat-db")

# Ten file -> ma market, KHOP voi ma dung trong shopee_db.market_from_link() (vd
# 'cat-id.xlsx' la thi truong Indonesia 'id' - trung ten voi cot 'cat_id' nhung KHAC nghia,
# de y khi doc).
_FILE_BY_MARKET = {
    "ph": "cat-ph.xlsx",
    "th": "cat-th.xlsx",
    "my": "cat-my.xlsx",
    "id": "cat-id.xlsx",
    "vn": "cat-vn.xlsx",
    "sg": "cat-sg.xlsx",
}

_cache = None


def _load_one(path):
    """{cat_id(int): cat_name(str)} tu 1 file xlsx - bo qua file thieu/sai cot thay vi loi
    tran (du lieu tham khao, khong duoc lam gian doan luong import root chinh)."""
    names = {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except (OSError, ValueError):
        return names
    try:
        ws = wb["Shopee_Data"] if "Shopee_Data" in wb.sheetnames else wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return names
        col_index = {str(h).strip().lower(): i for i, h in enumerate(header) if h}
        cat_id_i = col_index.get("cat_id")
        cat_name_i = col_index.get("cat_name")
        if cat_id_i is None or cat_name_i is None:
            return names
        for row in rows:
            if row is None or len(row) <= max(cat_id_i, cat_name_i):
                continue
            raw_id = row[cat_id_i]
            if raw_id is None:
                continue
            try:
                cat_id = int(raw_id)
            except (TypeError, ValueError):
                continue
            raw_name = row[cat_name_i]
            names[cat_id] = str(raw_name).strip() if raw_name else None
    finally:
        wb.close()
    return names


def _load_all(cat_db_dir=CAT_DB_DIR):
    result = {}
    for market, filename in _FILE_BY_MARKET.items():
        path = os.path.join(cat_db_dir, filename)
        if os.path.isfile(path):
            result[market] = _load_one(path)
    return result


def _get_cache():
    global _cache
    if _cache is None:
        _cache = _load_all()
    return _cache


def cat_name_for(market, cat_id):
    """Tra ten category theo (market, cat_id) - None neu khong xac dinh duoc (thieu market,
    market khong co trong cat-db, hoac cat_id khong khop dong nao - Shopee co the da them
    category moi sau lan xuat file cat-db gan nhat, van giu cat_id, chi thieu ten hien thi)."""
    if not market or cat_id in (None, ""):
        return None
    try:
        cat_id_int = int(cat_id)
    except (TypeError, ValueError):
        return None
    return _get_cache().get(market, {}).get(cat_id_int)
