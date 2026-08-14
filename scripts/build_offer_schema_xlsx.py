"""Lay TIEU DE COT truc tiep tu schema SQL that cua database
(D:\\Shopee369\\0-database\\migrate\\schema.sql) - KHONG lay/them bat cu gi tu du lieu
JSON da capture trong du an nay. Moi sheet = 1 bang, hang 1 = dung ten cot, dung thu tu,
theo dinh nghia CREATE TABLE trong file schema.sql (parse truc tiep, khong go tay).

Chay:
    python scripts/build_offer_schema_xlsx.py
    python scripts/build_offer_schema_xlsx.py --schema-sql "D:/Shopee369/0-database/migrate/schema.sql" --output "D:/Shopee369/0-database/affiliate_offer_schema.xlsx"
"""
import argparse
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SCHEMA_SQL_DEFAULT = "D:/Shopee369/0-database/migrate/schema.sql"
OUTPUT_DEFAULT = "D:/Shopee369/0-database/affiliate_offer_schema.xlsx"

CREATE_TABLE_RE = re.compile(
    r"create table\s+public\.(\w+)\s*\((.*?)\n\)\s*;",
    re.IGNORECASE | re.DOTALL,
)


def parse_tables(sql_text):
    """Tra ve danh sach (ten_bang, [ten_cot, ...]) theo dung thu tu trong file, bang
    cach parse truc tiep cac khoi 'create table public.<ten> ( ... );'."""
    tables = []
    for match in CREATE_TABLE_RE.finditer(sql_text):
        table_name = match.group(1)
        body = match.group(2)
        columns = []
        for raw_line in body.splitlines():
            line = raw_line.strip().rstrip(",")
            if not line:
                continue
            col_name = line.split()[0]
            columns.append(col_name)
        tables.append((table_name, columns))
    return tables


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--schema-sql", default=SCHEMA_SQL_DEFAULT)
    ap.add_argument("--output", default=OUTPUT_DEFAULT)
    args = ap.parse_args()

    with open(args.schema_sql, encoding="utf-8") as f:
        sql_text = f.read()

    tables = parse_tables(sql_text)
    if not tables:
        print(f"Khong tim thay 'create table public.<ten> (...)' nao trong {args.schema_sql}")
        return

    wb = Workbook()
    wb.remove(wb.active)
    for table_name, columns in tables:
        ws = wb.create_sheet(table_name)
        ws.append(columns)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
        for i in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22
        print(f"Sheet '{table_name}': {len(columns)} cot -> {columns}")

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    wb.save(args.output)
    print(f"Da ghi {args.output} ({len(tables)} sheet, tu {args.schema_sql})")


if __name__ == "__main__":
    main()
