"""Local HTTP server - cau noi giua Tampermonkey (chay trong tung Chrome profile, tu goi
API affiliate.shopee.* that qua fetch() cua trang, co san cookie dang nhap) va SQLite
(shopee_db.py). Ly do can server rieng: Tampermonkey/trinh duyet khong doc/ghi SQLite truc
tiep duoc; nhieu Chrome profile chay SONG SONG (nhieu tai khoan) can 1 noi TRUNG TAM giu
tinh nguyen tu khi gan item vao group (try_assign_verified) de khong bi 2 profile gianh
trung 1 san pham cho 2 group khac nhau.

Chi bind 127.0.0.1 (khong 0.0.0.0) - server nay khong danh cho truy cap tu may khac.
Tampermonkey goi qua GM_xmlhttpRequest (khong phai fetch() thuong cua trang) de ne CORS
hoan toan - server KHONG can bat CORS.

Chay:
    python scripts/affiliate_scrape_server.py
    python scripts/affiliate_scrape_server.py --port 8877 --db-path artifacts/db/shopee.db
"""
import argparse
import io
import os
import re
import shutil
import socket
import subprocess
import sys
import threading
import time
from datetime import datetime

from flask import Flask, Response, abort, jsonify, render_template, request
from openpyxl import Workbook, load_workbook

import chrome_launcher
import dongvanfb_client
import gsheet_push_api
import microsoft_mail_client
import shopee_categories
import shopee_db
import videoai_client

app = Flask(__name__)
gsheet_push_api.register(app)  # tab "Push Sheet" - xem scripts/gsheet_push_api.py
DB_PATH = shopee_db.DB_PATH_DEFAULT  # ghi de qua --db-path luc khoi dong, xem main()
LAUNCH_URL_DEFAULT = "https://affiliate.shopee.ph/offer/product_offer"
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
USERSCRIPTS_DIR = os.path.join(SCRIPTS_DIR, "userscripts")
REPO_ROOT = os.path.dirname(SCRIPTS_DIR)  # thu muc goc git (chua .git/) - dung cho /api/update/*
UPDATE_RESTART_EXIT_CODE = 42  # start_affiliate_scraper.bat doc ma nay de tu khoi dong lai

# Nguon chan ly DUY NHAT cho moi thu lien quan userscript - dashboard (index.html) doc
# qua GET /api/userscripts thay vi hardcode ten/mo ta rieng, tranh 2 noi bi lech nhau.
# Cung la danh sach trang cho /userscripts/<file> (KHONG phuc vu file ngoai danh sach nay,
# du server chi bind 127.0.0.1).
USERSCRIPTS = [
    {
        "file": "tampermonkey_affiliate_group_scraper.user.js",
        "title": "Affiliate Offer Group Scraper",
        "description": "Script CHÍNH: Chạy trên affiliate.shopee.* để gom nhóm 6 sản phẩm/group (root đạt chuẩn + tối đa 5 sản phẩm tương tự). Cần cho MỖI tài khoản/profile đang dùng để cào.",
    },
    {
        "file": "shopee_collector.user.js",
        "title": "Shopee Product Link Collector",
        "description": "Chạy trên trang Shopee thường (không phải affiliate) - cuộn trang tự động thu thập link sản phẩm, đẩy thẳng làm root vào DB hoặc xuất TXT/JSON/CSV.",
    },
    {
        "file": "shopee_ph_phone_checker.user.js",
        "title": "Shopee PH Phone Checker (SMSPool + 5sim + dongvanfb Mail)",
        "description": "1 script, 2 vai trò theo domain đang mở: Trên bất kỳ trang nào của shopee.ph - lấy số PH từ SMSPool/5sim qua API key hoặc dongvanfb mail, kiểm tra check_phone_exist, tự hủy số đã tồn tại; Trên 5sim.net - mua/hủy số bằng chính session trình duyệt (không qua API key, né rate limit riêng), gửi yêu cầu check sang tab shopee.ph qua GM_addValueChangeListener rồi tự quyết định hủy/giữ.",
    },
    {
        "file": "tampermonkey_affiliate_root_navigator.user.js",
        "title": "Affiliate Root Navigator (Navigation) - bản chống Page Unavailable",
        "description": "THAY THẾ group scraper cũ khi Shopee chặn gọi offer/product liên tiếp (Page Unavailable sau ~1 root). Shopee chỉ chấp nhận token 'af-ac-enc-sz-token' mint từ 1 report (df.infra) và mỗi report chỉ dùng được ĐÚNG 1 lần - chỉ load trang thật mới kích engine gửi report. Script điều hướng tab tới offer/product_offer/<item_id> cho TỪNG root (như người mở link), hook fetch từ document-start để chụp response offer do CHÍNH TRANG gọi (token hợp lệ), rồi đẩy server local xử lý verify/seed/gan group/finish - KHÔNG gọi thêm request thật nào tới Shopee. Cách dùng: cài script này + TẮT script 'Affiliate Offer Group Scraper' cũ, nhập device key, Start. Mỗi root tốn 1 lần load trang (~3-8s) nhưng không bị chặn kiểu token reuse.",
    },
]
USERSCRIPT_ALLOWLIST = {u["file"] for u in USERSCRIPTS}

_VERSION_RE = re.compile(r"^//\s*@version\s+(\S+)", re.MULTILINE)


def _userscript_version(filename):
    """Doc truc tiep dong '// @version' tu file .user.js that (KHONG hardcode trong
    USERSCRIPTS - se le voi file that moi lan bump version). None neu khong doc duoc/khong
    co dong @version."""
    try:
        with open(os.path.join(USERSCRIPTS_DIR, filename), "r", encoding="utf-8") as f:
            content = f.read()
    except OSError:
        return None
    m = _VERSION_RE.search(content)
    return m.group(1) if m else None


@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")


@app.route("/api/userscripts", methods=["GET"])
def list_userscripts():
    items = [dict(u, version=_userscript_version(u["file"])) for u in USERSCRIPTS]
    return jsonify({"userscripts": items})


@app.route("/userscripts/<name>", methods=["GET"])
def userscript(name):
    """Phuc vu file .user.js THAT tu thu muc scripts/userscripts/ - dat @updateURL/
    @downloadURL trong header script tro ve day de Tampermonkey TU phat hien ban moi (so
    @version) va hien man hinh "Update" - khong can copy/dan tay nua. Content-Type dung
    de Tampermonkey/trinh duyet nhan dien day la userscript."""
    if name not in USERSCRIPT_ALLOWLIST:
        abort(404)
    path = os.path.join(USERSCRIPTS_DIR, name)
    with open(path, "r", encoding="utf-8") as f:
        content = f.read()
    return Response(content, content_type="text/javascript; charset=utf-8")


def _run_git(args, timeout=30):
    """Chay 1 lenh git trong REPO_ROOT (KHONG dua vao cwd cua tien trinh - server co the
    duoc khoi dong tu bat ky thu muc nao). Tra ve subprocess.CompletedProcess (khong raise
    khi git tra ma loi khac 0 - nguoi goi tu kiem tra returncode)."""
    return subprocess.run(
        ["git"] + args,
        cwd=REPO_ROOT,
        capture_output=True,
        text=True,
        timeout=timeout,
    )


@app.route("/api/update/check", methods=["GET"])
def update_check():
    """Kiem tra co ban code moi tren remote 'origin' khong (git fetch + so sanh HEAD local
    voi origin/master) - dung cho nut 'Kiem tra cap nhat' tren dashboard. Repo PUBLIC
    (SHOPEE_AFF_AIO) nen fetch khong can dang nhap/token gi ca."""
    fetch = _run_git(["fetch", "origin", "master"])
    if fetch.returncode != 0:
        return jsonify({"error": f"git fetch that bai: {fetch.stderr.strip() or fetch.stdout.strip()}"}), 500

    local = _run_git(["rev-parse", "HEAD"])
    remote = _run_git(["rev-parse", "origin/master"])
    if local.returncode != 0 or remote.returncode != 0:
        return jsonify({"error": "Khong doc duoc commit hien tai - thu muc nay co phai git repo (da git clone) khong?"}), 500

    local_sha = local.stdout.strip()
    remote_sha = remote.stdout.strip()
    update_available = local_sha != remote_sha

    commits = []
    if update_available:
        log = _run_git(["log", f"{local_sha}..{remote_sha}", "--pretty=format:%h %s"])
        commits = [line for line in log.stdout.splitlines() if line.strip()]

    return jsonify({
        "update_available": update_available,
        "local_commit": local_sha[:7],
        "remote_commit": remote_sha[:7],
        "commits": commits,
    })


@app.route("/api/update/apply", methods=["POST"])
def update_apply():
    """Tai ban code moi nhat (git pull --ff-only tu origin/master) roi TU KHOI DONG LAI -
    dung cho nut 'Cap nhat ngay'. --ff-only: tu choi neu co local change/lich su re nhanh
    (an toan - KHONG bao gio tao merge commit hay ghi de am tham), tra loi ro nguyen nhan
    thay vi lam hong thu muc lam viec.

    Tu restart bang cach thoat tien trinh voi UPDATE_RESTART_EXIT_CODE (42) - script khoi
    dong (start_affiliate_scraper.bat) doc ma nay va TU chay lai python, ap dung code vua
    pull. Tra response VE TRUOC (qua thread nen + sleep ngan) de trinh duyet nhan duoc ket
    qua truoc khi tien trinh bi os._exit() (ngat ngang, khong chay cleanup) - day la ly do
    can thread rieng thay vi goi os._exit() ngay tai day."""
    pull = _run_git(["pull", "--ff-only", "origin", "master"])
    if pull.returncode != 0:
        return jsonify({"error": f"git pull that bai: {pull.stderr.strip() or pull.stdout.strip()}"}), 500

    def _restart_soon():
        time.sleep(1)
        os._exit(UPDATE_RESTART_EXIT_CODE)

    threading.Thread(target=_restart_soon, daemon=True).start()
    return jsonify({
        "ok": True,
        "output": pull.stdout.strip(),
        "message": "Da cap nhat code moi nhat - server dang tu khoi dong lai...",
    })


def _bad_request(msg):
    return jsonify({"error": msg}), 400


@app.errorhandler(Exception)
def _handle_error(e):
    # Khong de loi tran ra thanh HTML mac dinh cua Flask - Tampermonkey/dashboard doc
    # JSON. QUAN TRONG: HTTPException (vd tu abort(404)) da co san status code dung -
    # phai giu nguyen, khong de tat ca roi xuong 500 (da gap bug that: abort(404) o
    # /userscripts/<name> bi handler nay de thanh 500).
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return jsonify({"error": e.description}), e.code
    return jsonify({"error": str(e)}), 500


def _parse_cat_id(value):
    if value in (None, ""):
        return None
    return int(value)


@app.route("/api/roots/import", methods=["POST"])
def import_roots():
    body = request.get_json(force=True, silent=True) or {}
    links = body.get("links")
    if not isinstance(links, list) or not links:
        return _bad_request("thieu 'links' (danh sach string)")

    # cat_ids: cat_id RIENG cho tung link (danh sach song song voi links) - Shopee Product
    # Link Collector gui theo dinh dang nay khi cao nhieu danh muc truoc khi day vao DB 1
    # lan, tranh gan sai cat_id cuoi cung cho toan bo lo (xem shopee_db.import_roots_as_pending).
    cat_ids_raw = body.get("cat_ids")
    if cat_ids_raw is not None:
        if not isinstance(cat_ids_raw, list) or len(cat_ids_raw) != len(links):
            return _bad_request("'cat_ids' phai la danh sach cung do dai voi 'links'")
        try:
            cat_ids = [_parse_cat_id(v) for v in cat_ids_raw]
        except (TypeError, ValueError):
            return _bad_request("'cat_ids' chi duoc chua so nguyen hoac null")
        added = shopee_db.import_roots_as_pending(DB_PATH, links, cat_ids=cat_ids)
        return jsonify({"added": added})

    try:
        cat_id = _parse_cat_id(body.get("cat_id"))
    except (TypeError, ValueError):
        return _bad_request("'cat_id' phai la so nguyen")
    added = shopee_db.import_roots_as_pending(DB_PATH, links, cat_id=cat_id)
    return jsonify({"added": added})


@app.route("/api/roots/claim", methods=["POST"])
def claim_root():
    body = request.get_json(force=True, silent=True) or {}
    device_key = body.get("device_key")
    market = body.get("market")
    if not device_key:
        return _bad_request("thieu 'device_key' (ten tai khoan/profile dang claim)")
    if not market:
        return _bad_request("thieu 'market' (tab chi duoc claim root DUNG market no dang mo)")
    row = shopee_db.claim_root(DB_PATH, device_key, market)
    return jsonify({"root": row})


@app.route("/api/roots/<itemid>/assign", methods=["POST"])
def assign_root(itemid):
    body = request.get_json(force=True, silent=True) or {}
    device_key = body.get("device_key")
    market = body.get("market")
    if not device_key:
        return _bad_request("thieu 'device_key'")
    if not market:
        return _bad_request("thieu 'market'")
    result = shopee_db.assign_root_to_worker(DB_PATH, itemid, device_key, market)
    return jsonify(result), (200 if result["ok"] else 409)


@app.route("/api/workers/<device_key>/assigned_root", methods=["GET"])
def assigned_root(device_key):
    market = request.args.get("market")
    if not market:
        return _bad_request("thieu 'market' (tab chi duoc giao root DUNG market no dang mo)")
    root = shopee_db.get_assigned_root_for_worker(DB_PATH, device_key, market)
    return jsonify({"root": root})


@app.route("/api/workers/heartbeat", methods=["POST"])
def workers_heartbeat():
    body = request.get_json(force=True, silent=True) or {}
    device_key = body.get("device_key")
    status = body.get("status")
    if not device_key or not status:
        return _bad_request("thieu 'device_key' hoac 'status'")
    shopee_db.worker_heartbeat(DB_PATH, device_key, status, body.get("current_root"), body.get("market"))
    return jsonify({"ok": True})


@app.route("/api/workers", methods=["GET"])
def workers_list():
    return jsonify({"workers": shopee_db.list_workers(DB_PATH)})


@app.route("/api/workers/<device_key>", methods=["DELETE"])
def remove_worker(device_key):
    """Xoa 1 device_key khoi bang 'workers' - dung cho nut 'Xoa' o tab 'Van hanh'. Nha kem
    claim (assigned_key) cua device_key nay tren products, xem shopee_db.remove_worker()."""
    result = shopee_db.remove_worker(DB_PATH, device_key)
    return jsonify(result)


@app.route("/api/roots/<itemid>/reset", methods=["POST"])
def reset_root(itemid):
    body = request.get_json(force=True, silent=True) or {}
    market = body.get("market")
    if not market:
        return _bad_request("thieu 'market'")
    ok = shopee_db.reset_root_to_pending(DB_PATH, itemid, market)
    if not ok:
        return _bad_request(f"khong tim thay root '{itemid}' o market '{market}'")
    return jsonify({"ok": True})


@app.route("/api/roots/<itemid>/fail", methods=["POST"])
def fail_root(itemid):
    """Worker goi khi API tra loi that su cho chinh root (vd 'invalid item id') - nha
    claim + chuyen status_link='fail' de KHONG bi nhan lai vo han lan sau."""
    body = request.get_json(force=True, silent=True) or {}
    reason = body.get("reason") or "unknown_error"
    market = body.get("market")
    if not market:
        return _bad_request("thieu 'market'")
    shopee_db.mark_root_failed(DB_PATH, itemid, market, reason)
    return jsonify({"ok": True})


@app.route("/api/roots/<itemid>/recompute_merged", methods=["POST"])
def recompute_merged(itemid):
    """Tinh lai merged_link thu cong - dung cho group da 'done' TU TRUOC KHI co tinh nang
    nay (finish_root() tu dong lam viec nay cho group hoan tat SAU nay)."""
    body = request.get_json(force=True, silent=True) or {}
    market = body.get("market")
    if not market:
        return _bad_request("thieu 'market'")
    total = shopee_db.compute_merged_links(DB_PATH, itemid, market)
    return jsonify({"ok": True, "total_links": total})


@app.route("/api/roots/recompute_merged_all", methods=["POST"])
def recompute_merged_all():
    result = shopee_db.recompute_all_merged_links(DB_PATH)
    return jsonify(result)


@app.route("/api/roots/reset_insufficient_all", methods=["POST"])
def reset_insufficient_all():
    result = shopee_db.reset_all_insufficient_roots(DB_PATH)
    return jsonify(result)


@app.route("/api/candidates/recheck_cached", methods=["POST"])
def recheck_cached_candidates():
    """1 lan bam nut = 2 buoc: (1) giai phong related cua cac root DA TUNG dat nhung KHONG
    CON dat dieu kien HIEN TAI (release_disqualified_root_members() - dam bao dung nguyen
    tac "root khong dat thi khong duoc giu related" MOI LUC, khong chi luc cao lan dau),
    (2) quet lai TOAN BO candidate 'cached' (gom ca vua giai phong o buoc 1) doi chieu voi
    dieu kien hien tai va gan lai cho root phu hop neu co (recheck_cached_candidates()).
    Ca 2 buoc KHONG goi API Shopee. body (optional): {"market": "ph"} de gioi han 1 thi
    truong."""
    body = request.get_json(force=True, silent=True) or {}
    market = body.get("market") or None
    release_result = shopee_db.release_disqualified_root_members(DB_PATH, market=market)
    recheck_result = shopee_db.recheck_cached_candidates(DB_PATH, market=market)
    return jsonify({
        "roots_disqualified": release_result["roots_disqualified"],
        "released": len(release_result["released_itemids"]),
        "checked": recheck_result["checked"],
        "assigned": recheck_result["assigned"],
    })


@app.route("/api/roots/nav_complete", methods=["POST"])
def nav_complete():
    """1 lan goi duy nhat cho 1 root o che do "Root Navigator" (userscript dieu huong trang
    that): server nhan offer_data MA CHINH TRANG Shopee da goi (token da hop le), tu verify
    root, neu DAT thi seed + gan related (toi 5) + finish - gom toan bo logic truoc day
    userscript phai goi nhieu lan (verify/seed/items.verify/finish) thanh 1 request local duy
    nhat, giam diem loi va round-trip. KHONG goi bat ky API Shopee nao o day."""
    body = request.get_json(force=True, silent=True) or {}
    offer_data = body.get("offer_data")
    if not isinstance(offer_data, dict):
        return _bad_request("thieu 'offer_data' (object response.data cua offer/product)")
    itemid = str(offer_data.get("item_id") or "")
    market = body.get("market") or shopee_db.market_from_link(offer_data.get("product_link"))
    if not itemid or not market:
        return _bad_request("khong suy duoc itemid/market tu offer_data")
    row = shopee_db.map_v2_data_to_row(
        offer_data, link_type="root", groupid=itemid, market=market
    )
    verify = shopee_db.verify_root(DB_PATH, offer_data)
    if not verify.get("passes"):
        shopee_db.finish_root(DB_PATH, itemid, market)
        return jsonify({"ok": True, "outcome": "rejected", "itemid": itemid})

    settings = shopee_db.get_settings(DB_PATH)
    sold_min = settings.get("sold_min") or 0
    similar = (offer_data.get("similar_product_offers") or {}).get("list") or []
    claimed = shopee_db.seed_and_claim_candidates(DB_PATH, itemid, similar, market=market) or []
    claimed_set = {str(c) for c in claimed}
    candidates = []
    for it in similar:
        sid = str(it.get("item_id") or "")
        if not sid or sid == itemid or sid not in claimed_set:
            continue
        try:
            sold = int((it.get("batch_item_for_item_card_full") or {}).get("sold") or 0)
        except (TypeError, ValueError):
            sold = 0
        if sold <= sold_min:
            continue
        candidates.append((sold, it))
    candidates.sort(key=lambda c: -c[0])

    member = 0
    errors = []
    detail = {"similar_total": len(similar), "claimed": len(claimed), "sold_passed": 0,
              "outcomes": {"assigned": 0, "already_member": 0, "failed_criteria": 0,
                           "claimed_by_other": 0, "error": 0}}
    for sold, it in candidates:
        if member >= 5:  # GROUP_TARGET-1
            break
        detail["sold_passed"] += 1
        try:
            related_row = shopee_db.map_v2_data_to_row(
                it, link_type="related", groupid=itemid, market=market
            )
            if not related_row.get("itemid"):
                continue
            # Candidate chi co COMMISSION THEO TY LE % (seller_commission_rate/default_commission_rate)
            # ma khong kem so tien (da xac nhan that 2026-09-03: similar_product_offers.list chi
            # tra rate, cr=null) - tieu chi hien tai can so tien. Uoc luong: pct% * gia hien thi
            # (price int / 100000 = gia hien thi PHP/TH, vi du 19900000 -> ₱199.00) de co so sanh.
            # Neu response da co so tien that (commission_rate.seller_commission) thi giu nguyen.
            if not related_row.get("seller_commission"):
                it_batch = it.get("batch_item_for_item_card_full") or {}
                pct_raw = it.get("seller_commission_rate") or it.get("default_commission_rate")
                try:
                    price_raw = int(it_batch.get("price") or 0)
                except (TypeError, ValueError):
                    price_raw = 0
                if pct_raw and price_raw:
                    try:
                        pct = float(str(pct_raw).replace("%", "").strip()) / 100.0
                        price_display = price_raw / 100000.0 if price_raw > 100000 else float(price_raw)
                        est = round(pct * price_display, 2)
                        if est > 0:
                            related_row["seller_commission"] = est
                    except (TypeError, ValueError):
                        pass
            out = shopee_db.try_assign_verified(DB_PATH, related_row, itemid)
            if out:
                oc = out.get("outcome")
                if oc in detail["outcomes"]:
                    detail["outcomes"][oc] += 1
            if out and out.get("outcome") in ("assigned", "already_member"):
                member = out.get("group_member_count") or member
        except Exception as e:  # noqa: BLE001 - 1 candidate loi khong duoc lam chet ca root
            errors.append({"itemid": it.get("item_id"), "error": str(e)[:200]})
            detail["outcomes"]["error"] += 1
    shopee_db.finish_root(DB_PATH, itemid, market)
    return jsonify({
        "ok": True,
        "outcome": "done",
        "itemid": itemid,
        "member_count": member,
        "errors": errors[:20],
        "detail": detail,
    })


@app.route("/api/roots/finish", methods=["POST"])
def finish_root():
    body = request.get_json(force=True, silent=True) or {}
    itemid = body.get("itemid")
    market = body.get("market")
    if not itemid:
        return _bad_request("thieu 'itemid'")
    if not market:
        return _bad_request("thieu 'market'")
    shopee_db.finish_root(DB_PATH, itemid, market)
    return jsonify({"ok": True})


@app.route("/api/candidates/seed", methods=["POST"])
def seed_candidates():
    """items: nguyen si similar_product_offers.list[] tu response Shopee. Tra ve
    'claimed_item_ids' - CHI cac item_id nhom nay thuc su duoc giu (moi hoac da la cua
    minh tu truoc); item da bi nhom khac giu se KHONG co trong danh sach - BFS phia goi
    dung danh sach nay de biet item nao dang duoc xep vao hang doi cua chinh minh."""
    body = request.get_json(force=True, silent=True) or {}
    groupid = body.get("groupid")
    items = body.get("items")
    if not groupid or not isinstance(items, list):
        return _bad_request("thieu 'groupid' hoac 'items' (danh sach)")
    claimed = shopee_db.seed_and_claim_candidates(DB_PATH, groupid, items)
    return jsonify({"claimed_item_ids": claimed})


@app.route("/api/roots/verify", methods=["POST"])
def verify_root():
    """offer_data: nguyen si response.data tu goi that offer/product?item_id=<root>. Chi
    cap nhat metrics that cho dong root (KHONG doi status_link/claim). Tra ve 'passes' de
    userscript quyet dinh: KHONG dat -> loai luon (khong lay san pham tuong tu); DAT -> lay
    toi da 5 san pham tuong tu tu chinh similar_product_offers cua root cho du nhom 6."""
    body = request.get_json(force=True, silent=True) or {}
    offer_data = body.get("offer_data")
    if not isinstance(offer_data, dict):
        return _bad_request("thieu 'offer_data' (object)")
    result = shopee_db.verify_root(DB_PATH, offer_data)
    return jsonify(result)


@app.route("/api/items/verify", methods=["POST"])
def verify_item():
    """offer_data: nguyen si response.data tu goi that offer/product?item_id=<candidate> -
    tuc DA ton 1 request that toi Shopee cho item nay. Server tinh tieu chi + gan group
    (nguyen tu, an toan khi nhieu profile goi song song)."""
    body = request.get_json(force=True, silent=True) or {}
    groupid = body.get("groupid")
    offer_data = body.get("offer_data")
    if not groupid or not isinstance(offer_data, dict):
        return _bad_request("thieu 'groupid' hoac 'offer_data' (object)")
    row = shopee_db.map_v2_data_to_row(offer_data, link_type="related", groupid=groupid)
    if not row.get("itemid"):
        return _bad_request("offer_data khong co item_id hop le")
    result = shopee_db.try_assign_verified(DB_PATH, row, groupid)
    return jsonify(result)


@app.route("/api/items/filter_new", methods=["POST"])
def filter_new():
    body = request.get_json(force=True, silent=True) or {}
    itemids = body.get("itemids")
    if not isinstance(itemids, list):
        return _bad_request("thieu 'itemids' (danh sach)")
    new_ids = shopee_db.filter_new_itemids(DB_PATH, itemids)
    return jsonify({"new_itemids": new_ids})


@app.route("/api/items/list", methods=["GET"])
def list_items():
    """Danh sach san pham da cao (tab 'San pham' tren dashboard) - loc theo
    link_type/status_link/search (khop ten cot itemid/name/shop_name)/groupid (khop chinh
    xac 1 nhom), gioi han so dong. status_link='video_ready' = loc 'du dieu kien tao video'
    (root co merged_link, chua tao job VideoAI, co product_link - cung dieu kien voi hang
    doi tao video). Tra kem 'total' = tong so link khop bo loc (khong gioi han)."""
    link_type = request.args.get("link_type") or None
    status_link = request.args.get("status_link") or None
    search = request.args.get("search") or None
    groupid = request.args.get("groupid") or None
    market = request.args.get("market") or None
    limit = min(max(1, request.args.get("limit", 200, type=int)), 500)
    if status_link == "video_ready":
        items, total = shopee_db.video_ready_items(
            DB_PATH, market=market, search=search, groupid=groupid, limit=limit
        )
        return jsonify({"items": items, "total": total, "mode": "video_ready"})
    items = shopee_db.fetch_all_items(
        DB_PATH, link_type=link_type, status_link=status_link, search=search,
        groupid=groupid, market=market, limit=limit,
    )
    total = shopee_db.count_all_items(
        DB_PATH, link_type=link_type, status_link=status_link, search=search,
        groupid=groupid, market=market,
    )
    return jsonify({"items": items, "total": total})


@app.route("/api/items/<itemid>", methods=["DELETE"])
def delete_item(itemid):
    market = request.args.get("market")
    if not market:
        return _bad_request("thieu query param 'market'")
    ok = shopee_db.delete_item(DB_PATH, itemid, market)
    if not ok:
        return _bad_request(f"khong tim thay item '{itemid}' o market '{market}'")
    return jsonify({"ok": True})


@app.route("/api/groups/<groupid>/count", methods=["GET"])
def group_count(groupid):
    market = request.args.get("market")
    if not market:
        return _bad_request("thieu query param 'market'")
    count = shopee_db.count_group_members(DB_PATH, groupid, market)
    return jsonify({"groupid": groupid, "member_count": count})


@app.route("/api/roots/list", methods=["GET"])
def list_roots():
    market = request.args.get("market") or None
    status = request.args.get("status") or None
    return jsonify({"roots": shopee_db.list_roots_with_counts(DB_PATH, status=status, market=market)})


@app.route("/api/roots/reset_by_filter", methods=["POST"])
def reset_roots_by_filter():
    """Dat lai (ve 'pending') toan bo root theo bo loc dang chon o UI (market + trang thai) -
    dung cho nut 'Dat lai Root (theo bo loc)' o khoi 'Danh sach Root'. Chi cho phep reset root
    dang 'done'/'fail' (root pending khong co gi de reset). Nha claim + xoa fail_reason."""
    body = request.get_json(force=True, silent=True) or {}
    market = (body.get("market") or "").strip() or None
    status = (body.get("status") or "").strip()
    if status not in ("done", "fail"):
        return _bad_request("'status' chi ho tro 'done' hoac 'fail'")
    count = shopee_db.reset_roots_by_filter(DB_PATH, market=market, statuses=(status,))
    return jsonify({"ok": True, "reset_count": count, "market": market, "status": status})


@app.route("/api/items/category_stats", methods=["GET"])
def items_category_stats():
    market = request.args.get("market") or None
    return jsonify(shopee_db.category_stats(DB_PATH, market=market))


@app.route("/api/categories/name", methods=["GET"])
def category_name():
    # Dung cho Shopee Product Link Collector (shopee_collector.user.js) - hien ten danh muc
    # (vd "Pets") thay vi chi cat_id tho tren panel, ngay luc dang cao. Nhan 'url' (trang
    # Shopee hien tai) thay vi 'market' truc tiep - suy market qua market_from_link() DUNG
    # HAM CHUNG voi phan import root, tranh trung logic map domain->market o phia client.
    url = request.args.get("url") or ""
    market = shopee_db.market_from_link(url)
    try:
        cat_id = int(request.args.get("cat_id")) if request.args.get("cat_id") not in (None, "") else None
    except (TypeError, ValueError):
        return _bad_request("'cat_id' phai la so nguyen")
    cat_name = shopee_categories.cat_name_for(market, cat_id) if cat_id is not None else None
    return jsonify({"market": market, "cat_name": cat_name})


@app.route("/api/categories/list", methods=["GET"])
def categories_list():
    """Danh sach danh muc cap 1 cua market (suy tu 'url', dung ham chung market_from_link()
    - xem category_name() o tren) - dung cho dropdown chon danh muc khi cao theo tu khoa o
    Shopee Product Link Collector (tranh link cao tu tu khoa bi "mo coi" khong co danh muc)."""
    url = request.args.get("url") or ""
    market = shopee_db.market_from_link(url)
    return jsonify({"market": market, "categories": shopee_categories.list_categories(market)})


@app.route("/api/roots/market_stats", methods=["GET"])
def roots_market_stats():
    """Tong hop so root theo tung market - dung cho bang "Root theo market" + dropdown
    chon market cho auto-assign o tab "Van hanh"."""
    return jsonify({"markets": shopee_db.count_roots_by_market(DB_PATH)})


@app.route("/api/roots/<groupid>/members", methods=["GET"])
def root_members(groupid):
    market = request.args.get("market")
    if not market:
        return _bad_request("thieu query param 'market'")
    return jsonify({"members": shopee_db.list_group_members(DB_PATH, groupid, market)})


@app.route("/api/accounts", methods=["GET"])
def list_accounts():
    return jsonify({"accounts": shopee_db.list_devices(DB_PATH)})


@app.route("/api/accounts", methods=["POST"])
def add_account():
    body = request.get_json(force=True, silent=True) or {}
    name = body.get("name")
    profile_path = body.get("profile_path")
    if not name or not profile_path:
        return _bad_request("thieu 'name' hoac 'profile_path'")
    shopee_db.add_device(DB_PATH, name, profile_path)
    return jsonify({"ok": True})


@app.route("/api/accounts/<path:profile_path>", methods=["DELETE"])
def remove_account(profile_path):
    shopee_db.remove_device(DB_PATH, profile_path)
    return jsonify({"ok": True})


@app.route("/api/accounts/<int:device_id>", methods=["PUT"])
def update_account(device_id):
    body = request.get_json(force=True, silent=True) or {}
    name = body.get("name")
    profile_path = body.get("profile_path")
    if not name or not profile_path:
        return _bad_request("thieu 'name' hoac 'profile_path'")
    shopee_db.update_device(DB_PATH, device_id, name, profile_path)
    return jsonify({"ok": True})


@app.route("/api/accounts/<name>/launch", methods=["POST"])
def launch_account(name):
    accounts = shopee_db.list_devices(DB_PATH)
    match = next((a for a in accounts if a["name"] == name), None)
    if not match:
        return _bad_request(f"khong tim thay tai khoan '{name}'")
    body = request.get_json(force=True, silent=True) or {}
    url = body.get("url") or LAUNCH_URL_DEFAULT
    try:
        proc = chrome_launcher.launch_profile(match["serial"], url)
    except (RuntimeError, ValueError) as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"ok": True, "pid": proc.pid})


@app.route("/api/settings", methods=["GET"])
def get_settings():
    return jsonify(shopee_db.get_settings(DB_PATH))


@app.route("/api/settings", methods=["POST"])
def update_settings():
    body = request.get_json(force=True, silent=True) or {}
    result = shopee_db.update_settings(
        DB_PATH,
        promoted_7d_max=body.get("promoted_7d_max"),
        sold_min=body.get("sold_min"),
        seller_commission_vnd_min=body.get("seller_commission_vnd_min"),
        auto_assign=body.get("auto_assign"),
        dongvanfb_api_key=body.get("dongvanfb_api_key"),
        auto_assign_market=body.get("auto_assign_market"),
    )
    return jsonify(result)


@app.route("/api/video_machines", methods=["GET"])
def list_video_machines():
    return jsonify({"machines": shopee_db.list_video_machines(DB_PATH)})


@app.route("/api/video_machines", methods=["POST"])
def add_video_machine():
    body = request.get_json(force=True, silent=True) or {}
    name = (body.get("name") or "").strip()
    api_key = (body.get("api_key") or "").strip()
    tag = (body.get("tag") or "").strip()
    pool = (body.get("pool") or "selfhostPool").strip()
    if not name or not api_key or not tag:
        return _bad_request("thieu 'name', 'api_key' hoac 'tag'")
    shopee_db.add_video_machine(DB_PATH, name, api_key, tag, pool)
    return jsonify({"ok": True})


@app.route("/api/video_machines/<int:machine_id>", methods=["DELETE"])
def remove_video_machine(machine_id):
    ok = shopee_db.remove_video_machine(DB_PATH, machine_id)
    if not ok:
        return _bad_request(f"khong tim thay may id={machine_id}")
    return jsonify({"ok": True})


@app.route("/api/video_machines/<int:machine_id>/toggle", methods=["POST"])
def toggle_video_machine(machine_id):
    body = request.get_json(force=True, silent=True) or {}
    enabled = body.get("enabled")
    if enabled is None:
        return _bad_request("thieu 'enabled' (true/false)")
    ok = shopee_db.set_video_machine_enabled(DB_PATH, machine_id, enabled)
    if not ok:
        return _bad_request(f"khong tim thay may id={machine_id}")
    return jsonify({"ok": True})


@app.route("/api/video_machines/<int:machine_id>/tag", methods=["POST"])
def update_video_machine_tag(machine_id):
    """Doi Tag (thu muc) cua 1 may tao video - dung cho nut 'Sua' canh 'Xoa' o tab 'Tao
    video'."""
    body = request.get_json(force=True, silent=True) or {}
    tag = (body.get("tag") or "").strip()
    if not tag:
        return _bad_request("thieu 'tag'")
    ok = shopee_db.set_video_machine_tag(DB_PATH, machine_id, tag)
    if not ok:
        return _bad_request(f"khong tim thay may id={machine_id}")
    return jsonify({"ok": True})


@app.route("/api/videos/stats", methods=["GET"])
def video_stats():
    return jsonify(shopee_db.count_video_push_stats(DB_PATH))


@app.route("/api/videos/stats_by_market", methods=["GET"])
def video_stats_by_market():
    """Thong ke link tao video theo TUNG market - dung cho bang "Link theo market" o tab
    "Tao video"."""
    return jsonify({"markets": shopee_db.count_video_push_stats_by_market(DB_PATH)})


@app.route("/api/videos/export.xlsx", methods=["GET"])
def export_videos_xlsx():
    """Xuat toan bo san pham DA TAO VIDEO (job_id khong null) ra .xlsx, cot A|B|C =
    itemId|Ten san pham|Link gop - dung cho nut 'Xuat Excel' tren tab 'Tao video'."""
    items = shopee_db.list_video_created_items(DB_PATH)
    wb = Workbook()
    ws = wb.active
    ws.title = "Da tao video"
    ws.append(["itemId", "Ten san pham", "Link gop"])
    for it in items:
        ws.append([it["itemid"], it["name"], it["merged_link"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"shopee_video_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.route("/api/videos/reset", methods=["POST"])
def reset_videos():
    """Dat lai trang thai 've cho tao video' cho san pham da tao xong (job_id khong null) -
    dung cho nut 'Dat lai trang thai' o tab 'Tao video'. market (query/body, optional): chi
    reset 1 thi truong dang chon tren dropdown; bo trong = tat ca."""
    body = request.get_json(force=True, silent=True) or {}
    market = body.get("market") or None
    count = shopee_db.reset_video_jobs(DB_PATH, market=market)
    return jsonify({"reset": count})


@app.route("/api/videos/push", methods=["POST"])
def push_videos():
    """1 lo (toi da 200, gioi han cua chinh VideoAI): day cache (cho dong chua
    cache_uploaded) + tao task video (cho dong da/vua co cache) cho toi da 'limit' san pham
    dang cho (list_video_push_candidates), dung API key/tag/pool cua 1 may tao video cu the
    (machine_id, xem quan ly may o tab 'Tao video'). Dashboard tu goi lap lai endpoint nay
    (xem runVideoPush() trong index.html) toi khi 'done'=0 de xu ly het hang doi - moi lan
    goi CHI xu ly 1 lo, giu request nhanh + co progress ro rang thay vi 1 request khong lo
    cho hang nghin san pham."""
    body = request.get_json(force=True, silent=True) or {}
    limit = min(max(1, int(body.get("limit") or 200)), videoai_client.BATCH_LIMIT)
    machine_id = body.get("machine_id")
    market = body.get("market") or None
    if not machine_id:
        return _bad_request("thieu 'machine_id' - chon 1 may tao video truoc khi chay.")

    machine = shopee_db.get_video_machine(DB_PATH, machine_id)
    if not machine:
        return _bad_request(f"khong tim thay may id={machine_id}")
    if not machine.get("enabled"):
        return _bad_request(f"may '{machine['name']}' dang tat - bat len truoc khi dung.")
    api_key = machine["api_key"]
    tag = machine["tag"]
    pool = machine["pool"] or "selfhostPool"

    candidates = shopee_db.list_video_push_candidates(DB_PATH, limit, market=market)
    if not candidates:
        return jsonify({"done": 0, "pushed": 0, "created": 0, "errors": []})

    # Gom theo market (cot 'market' gio da dang tin - moi insert path deu tu suy dung tu
    # domain link, xem shopee_db.market_from_link()) vi language/prefix anh phu thuoc
    # market, va API tao task chi nhan 1 'language' chung cho ca lo.
    by_market = {}
    for row in candidates:
        by_market.setdefault(row.get("market"), []).append(row)

    total_pushed = 0
    total_created = 0
    errors = []

    for market, rows in by_market.items():
        language = videoai_client.language_for_market(market)
        url_to_itemid = {r["product_link"]: r["itemid"] for r in rows}
        url_to_merged_link = {r["product_link"]: r.get("merged_link") for r in rows}
        ready_urls = [r["product_link"] for r in rows if r.get("cache_uploaded")]
        need_cache_rows = [r for r in rows if not r.get("cache_uploaded")]

        if need_cache_rows:
            items = []
            for r in need_cache_rows:
                item = videoai_client.build_cache_item(r, market)
                if item is None:
                    errors.append({"itemid": r["itemid"], "reason": "thieu du lieu bat buoc (ten/link)"})
                    continue
                items.append(item)
            if items:
                try:
                    result = videoai_client.push_cache_batch(items, api_key)
                except Exception as e:
                    errors.append({"reason": f"loi day cache ({market}): {e}"})
                else:
                    failed_urls = {e.get("url") for e in (result.get("errors") or [])}
                    ok_itemids = []
                    for it in items:
                        if it["url"] in failed_urls:
                            errors.append({"itemid": url_to_itemid.get(it["url"]), "reason": "day cache that bai"})
                            continue
                        ok_itemids.append(url_to_itemid[it["url"]])
                        ready_urls.append(it["url"])
                    if ok_itemids:
                        shopee_db.mark_cache_uploaded(DB_PATH, [(iid, market) for iid in ok_itemids])
                        total_pushed += len(ok_itemids)

        if ready_urls:
            ready_items = [
                {"url": u, "merged_link": url_to_merged_link.get(u)} for u in ready_urls
            ]
            try:
                task_results = videoai_client.create_video_batch(
                    ready_items, api_key, tag=tag, pool=pool, language=language
                )
            except Exception as e:
                errors.append({"reason": f"loi tao video ({market}): {e}"})
            else:
                job_updates = []
                for it in task_results:
                    itemid = url_to_itemid.get(it.get("url"))
                    if not itemid:
                        continue
                    if it.get("jobId"):
                        job_updates.append((itemid, market, it["jobId"]))
                        total_created += 1
                    else:
                        errors.append({"itemid": itemid, "reason": it.get("error") or "tao video that bai"})
                if job_updates:
                    shopee_db.mark_video_jobs(DB_PATH, job_updates)

    shopee_db.log_video_push(
        DB_PATH, market, machine_id, machine["name"], limit,
        len(candidates), total_pushed, total_created, errors,
    )
    return jsonify({
        "done": len(candidates),
        "pushed": total_pushed,
        "created": total_created,
        "errors": errors,
    })


@app.route("/api/videos/log", methods=["GET"])
def videos_log():
    market = request.args.get("market") or None
    try:
        limit = min(max(1, int(request.args.get("limit") or 50)), 500)
        offset = max(0, int(request.args.get("offset") or 0))
    except (TypeError, ValueError):
        return _bad_request("'limit'/'offset' phai la so nguyen")
    return jsonify(shopee_db.list_video_push_log(DB_PATH, market=market, limit=limit, offset=offset))


@app.route("/api/videos/log/clear", methods=["POST"])
def clear_videos_log():
    """Xoa lich su 'Nhat ky tao video' - dung cho nut 'Xoa nhat ky' o tab 'Tao video'. market
    (body, optional): chi xoa 1 thi truong dang chon tren dropdown; bo trong = xoa tat ca."""
    body = request.get_json(force=True, silent=True) or {}
    market = body.get("market") or None
    count = shopee_db.clear_video_push_log(DB_PATH, market=market)
    return jsonify({"deleted": count})


# ---- Tab "Tao tai khoan Shopee" (mua mail dongvanfb + doc code) ----

@app.route("/api/mail_accounts/account_types", methods=["GET"])
def mail_account_types():
    return jsonify({"account_types": dongvanfb_client.ACCOUNT_TYPES})


@app.route("/api/mail_accounts/balance", methods=["GET"])
def mail_accounts_balance():
    api_key = (shopee_db.get_settings(DB_PATH).get("dongvanfb_api_key") or "").strip()
    if not api_key:
        return _bad_request("chua cau hinh dongvanfb API key (o tab 'Tao tai khoan Shopee').")
    balance = dongvanfb_client.get_balance(api_key)
    return jsonify({"balance": balance})


@app.route("/api/mail_accounts/buy", methods=["POST"])
def mail_accounts_buy():
    api_key = (shopee_db.get_settings(DB_PATH).get("dongvanfb_api_key") or "").strip()
    if not api_key:
        return _bad_request("chua cau hinh dongvanfb API key (o tab 'Tao tai khoan Shopee').")
    body = request.get_json(force=True, silent=True) or {}
    account_type = str(body.get("account_type") or "")
    quantity = max(1, int(body.get("quantity") or 1))
    if not account_type:
        return _bad_request("thieu 'account_type'")
    result = dongvanfb_client.buy_mail(api_key, account_type, quantity)
    if not result or not result.get("status"):
        return _bad_request("Mua mail that bai: " + str(result.get("message") if result else "khong ro loi"))
    data = result.get("data") or {}
    lines = data.get("list_data") or []
    added = shopee_db.add_mail_accounts_from_buy(DB_PATH, lines, account_type, data.get("order_code"))
    return jsonify({
        "ok": True, "added": added,
        "total_amount": data.get("total_amount"), "balance": data.get("balance"),
    })


@app.route("/api/mail_accounts/add_manual", methods=["POST"])
def mail_accounts_add_manual():
    body = request.get_json(force=True, silent=True) or {}
    lines = str(body.get("lines_text") or "").splitlines()
    result = shopee_db.add_mail_accounts_manual(DB_PATH, lines)
    return jsonify({"ok": True, **result})


@app.route("/api/mail_accounts/list", methods=["GET"])
def mail_accounts_list():
    market = request.args.get("market") or None
    slot = request.args.get("slot") or None
    search = request.args.get("search") or None
    limit = request.args.get("limit", 500, type=int)
    rows = shopee_db.list_mail_accounts(DB_PATH, market=market, slot=slot, search=search, limit=limit)
    return jsonify({"accounts": rows})


@app.route("/api/mail_accounts/export.xlsx", methods=["GET"])
def export_mail_accounts_xlsx():
    """Xuat cac mail DA dung de tao tai khoan Shopee (co shopee_id) ra .xlsx - dung cho nut
    'Xuat Excel' tren tab 'Tao tai khoan Shopee'."""
    accounts = shopee_db.list_created_mail_accounts(DB_PATH)
    wb = Workbook()
    ws = wb.active
    ws.title = "Tai khoan Shopee"
    ws.append(["Full info", "Email", "PassEmail", "Shopee_id", "Device", "Profile", "Slot", "Market", "Shopee_code", "Thoi gian tao"])
    for a in accounts:
        ws.append([
            a["full_info"], a["email"], a["password"], a["shopee_id"],
            a["device"], a["profile"], a["slot"], a["market"], a["shopee_code"], a["created_at"],
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"shopee_accounts_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


_IMPORT_MAIL_ACCOUNTS_COLUMN_MAP = {
    "full info": "full_info",
    "shopee_id": "shopee_id",
    "device": "device",
    "profile": "profile",
    "slot": "slot",
    "market": "market",
    "shopee_code": "shopee_code",
}


@app.route("/api/mail_accounts/import_xlsx", methods=["POST"])
def import_mail_accounts_xlsx():
    """Nhap lai file .xlsx dung dinh dang cot cua export_mail_accounts_xlsx() (nut 'Import
    Excel' canh 'Xuat Excel' tren tab 'Tao tai khoan Shopee') - xem
    shopee_db.import_mail_accounts_from_rows(). Doc header theo TEN cot (khong phu thuoc thu
    tu) de khop voi _IMPORT_MAIL_ACCOUNTS_COLUMN_MAP, bo qua cot 'Email'/'PassEmail'/'Thoi
    gian tao' (khong can, xem docstring ham do)."""
    file = request.files.get("file")
    if file is None or not file.filename:
        return _bad_request("chua chon file .xlsx")
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return _bad_request(f"khong doc duoc file .xlsx: {e}")
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return _bad_request("file rong, khong co dong tieu de")
    col_index = {str(h).strip().lower(): i for i, h in enumerate(header) if h}
    if "full info" not in col_index:
        return _bad_request("thieu cot 'Full info' - file khong dung dinh dang xuat")
    rows = []
    for row in rows_iter:
        if row is None:
            continue
        row_dict = {}
        for header_name, field in _IMPORT_MAIL_ACCOUNTS_COLUMN_MAP.items():
            idx = col_index.get(header_name)
            if idx is not None and idx < len(row):
                row_dict[field] = row[idx]
        rows.append(row_dict)
    result = shopee_db.import_mail_accounts_from_rows(DB_PATH, rows)
    return jsonify({"ok": True, **result})


@app.route("/api/mail_accounts/check_shopee_id", methods=["GET"])
def mail_accounts_check_shopee_id():
    """Tra cac mail KHAC dang dung cung shopee_id nay - dung cho canh bao chong nhap trung
    (popup xac nhan) o tab 'Tạo tài khoản Shopee' truoc khi luu that su."""
    shopee_id = (request.args.get("shopee_id") or "").strip()
    exclude_id = request.args.get("exclude_id", type=int)
    accounts = shopee_db.find_mail_accounts_by_shopee_id(DB_PATH, shopee_id, exclude_id=exclude_id)
    return jsonify({"accounts": accounts})


@app.route("/api/mail_accounts/<int:account_id>", methods=["POST"])
def mail_accounts_update(account_id):
    body = request.get_json(force=True, silent=True) or {}
    row = shopee_db.update_mail_account_fields(
        DB_PATH, account_id,
        shopee_id=body.get("shopee_id"), device=body.get("device"),
        profile=body.get("profile"), slot=body.get("slot"), market=body.get("market"),
    )
    if row is None:
        return _bad_request(f"khong tim thay mail id={account_id}")
    return jsonify({"account": row})


@app.route("/api/mail_accounts/<int:account_id>", methods=["DELETE"])
def mail_accounts_delete(account_id):
    ok = shopee_db.delete_mail_account(DB_PATH, account_id)
    if not ok:
        return _bad_request(f"khong tim thay mail id={account_id}")
    return jsonify({"ok": True})


@app.route("/api/mail_accounts/clear", methods=["POST"])
def mail_accounts_clear():
    deleted = shopee_db.clear_mail_accounts(DB_PATH)
    return jsonify({"ok": True, "deleted": deleted})


@app.route("/api/mail_accounts/<int:account_id>/get_code", methods=["POST"])
def mail_accounts_get_code(account_id):
    row = shopee_db.get_mail_account(DB_PATH, account_id)
    if not row:
        return _bad_request(f"khong tim thay mail id={account_id}")
    # Doc TRUC TIEP qua Graph (microsoft_mail_client) la duong CHINH - khong ton phi/khong
    # phu thuoc dich vu ngoai. Fallback ve dongvanfb CHI khi duong truc tiep loi (vd token
    # nay co van de rieng voi cach goi cua ta) - dongvanfb co 2 nhanh Graph/IMAP rieng, da
    # tung ghi nhan thuc te co tai khoan 1 nhanh loi nhung nhanh kia van doc duoc (xem
    # dongvanfb_client.py). Van luu lai refresh_token moi (Microsoft cap kem moi lan goi) nhu
    # thoi quen an toan - KHONG bat buoc, da kiem chung token cu van dung duoc binh thuong
    # sau khi "bi thay" (xem ghi chu dau file microsoft_mail_client.py).
    try:
        code, note, new_refresh_token = microsoft_mail_client.fetch_shopee_code(row["refresh_token"], row["client_id"])
        if new_refresh_token and new_refresh_token != row["refresh_token"]:
            shopee_db.update_mail_account_refresh_token(DB_PATH, account_id, new_refresh_token)
    except microsoft_mail_client.MicrosoftMailError as e:
        code, note = dongvanfb_client.fetch_shopee_code(row["email"], row["refresh_token"], row["client_id"])
        note = f"[Graph trực tiếp lỗi: {e}] Fallback dongvanfb -> {note}"
    shopee_db.set_mail_account_code(DB_PATH, account_id, code)
    return jsonify({"code": code, "note": note})


@app.route("/api/mail_accounts/<int:account_id>/activate_login", methods=["POST"])
def mail_accounts_activate_login(account_id):
    """Doc mail 'co lan dang nhap moi' cua Shopee (tu info@security.shopee.<tld>), trich
    link kich hoat dang "https://<tld>.shp.ee/dlink/<code>", roi mo link do bang Chrome.
    Uu tien mo DUNG profile Shopee da dang ky (cot 'Thiết bị' o tab Mail Accounts, khop ten
    voi 1 dong trong 'devices' - hop ly hon vi link kich hoat thuong can dang nhap dung
    tai khoan Shopee lien quan) - NEU khong dien 'Thiết bị' hoac khong khop dong nao, fallback
    ve 1 cua so Chrome RIENG co dinh (chrome_launcher.launch_activation_link()) de KHONG
    dieu huong tab/cua so nguoi dung dang lam viec (xem ghi chu ham do). CHUA co fallback
    dongvanfb cho tinh nang nay (dongvanfb khong co endpoint rieng cho link kich hoat, chi
    co endpoint doc ma OTP)."""
    row = shopee_db.get_mail_account(DB_PATH, account_id)
    if not row:
        return _bad_request(f"khong tim thay mail id={account_id}")
    try:
        link, note, new_refresh_token = microsoft_mail_client.fetch_login_link(row["refresh_token"], row["client_id"])
        if new_refresh_token and new_refresh_token != row["refresh_token"]:
            shopee_db.update_mail_account_refresh_token(DB_PATH, account_id, new_refresh_token)
    except microsoft_mail_client.MicrosoftMailError as e:
        return jsonify({"error": f"Lỗi đọc mail: {e}"}), 500
    if not link:
        return jsonify({"link": None, "note": note})
    device_name = (row.get("device") or "").strip()
    matched_device = None
    if device_name:
        devices = shopee_db.list_devices(DB_PATH)
        matched_device = next((d for d in devices if d["name"] == device_name), None)
    try:
        if matched_device:
            proc = chrome_launcher.launch_profile(matched_device["serial"], link)
        else:
            proc = chrome_launcher.launch_activation_link(link)
    except (RuntimeError, ValueError) as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"link": link, "note": note, "pid": proc.pid, "profile": matched_device["name"] if matched_device else "ActivationLinks"})


@app.route("/api/reset", methods=["POST"])
def reset_all():
    """Xoa du lieu san pham (khong dong tai khoan/profile) - dung cho nut "Xoa toan bo du
    lieu" tren UI. market optional trong body: '' hoac thieu = xoa TAT CA (hanh vi cu), 1
    ma market cu the = CHI xoa dong cua thi truong do."""
    body = request.get_json(force=True, silent=True) or {}
    market = body.get("market") or None
    deleted = shopee_db.clear_all_items(DB_PATH, market=market)
    return jsonify({"ok": True, "deleted": deleted, "market": market})


@app.route("/api/stats", methods=["GET"])
def stats():
    video_stats = shopee_db.count_video_push_stats(DB_PATH)
    return jsonify({
        "root": {
            "pending": shopee_db.count_status(DB_PATH, link_type="root", status_link="pending"),
            "done": shopee_db.count_status(DB_PATH, link_type="root", status_link="done"),
            "fail": shopee_db.count_status(DB_PATH, link_type="root", status_link="fail"),
        },
        "related": {
            "pending": shopee_db.count_status(DB_PATH, link_type="related", status_link="pending"),
            "member": shopee_db.count_status(DB_PATH, link_type="related", status_link="member"),
            "cached": shopee_db.count_status(DB_PATH, link_type="related", status_link="cached"),
        },
        # "Root đủ điều kiện tạo video": root co merged_link (nhom da du) - KHONG giam sau khi
        # da tao video (job_id/cache_uploaded tinh rieng ben tab Tạo Video), xem count_video_push_stats().
        "root_video_eligible": video_stats.get("eligible") or 0,
        "total_items": shopee_db.count_items(DB_PATH),
    })


# ============================================================================
# Tab "Cào root AFF" - cao san pham root theo TU KHOA qua trang
# affiliate.shopee.*/offer/product_offer (xem Cao_root_aff.txt). Worker la
# cdp_keyword_worker.mjs (spawn tu tab Vận hành GPM, mode 'keyword'): claim tu khoa
# pending, dieu khien CHINH TRANG affiliate search theo tu khoa do (de trang tu goi
# /api/v3/offer/product/list voi token chong bot hop le), hook Network chup tung trang
# roi day ve /api/keywords/page_done. Server loc tieu chi theo CAU HINH CAO worker gui
# kem moi trang: sold_min + comm_money_min = so TIEN hoa hong uoc tinh toi thieu
# (seller_commission_rate% * gia hien thi - "seller_com quy doi thanh tien") + filter_types
# (danh dau Xtra). CAC GIA TRI NAY CHI AP DUNG KHI CAO (nhap o tab Vận hành GPM luc start
# worker), KHONG LUU khi import tu khoa. Item dat tieu chi insert root pending vao bang
# 'products' chung (link_type 'root', groupid=itemid) - link da ton tai o bat ky dau trong
# DB thi bo qua (dup_skipped).
# ============================================================================
KEYWORD_MARKETS = ("ph", "th", "my", "id", "vn", "sg")


def _parse_keyword_text(text):
    """Doc noi dung dan len: ho tro ca dang "moi dong 1 tu khoa" lan dang file nhom
    (tieu de nhom tren 1 dong, sau do { cac tu khoa } - xem keyword_PH.txt). Dang file
    nhom: chi giu dong NAM TRONG { }; tieu de nhom/dau ngoac tu bo. Neu ca text khong co
    { } thi moi dong khong rong la 1 tu khoa."""
    lines = [ln.strip() for ln in (text or "").splitlines() if ln.strip()]
    if not lines:
        return []
    if any(ln in ("{", "}") for ln in lines):
        out, in_block = [], False
        for ln in lines:
            if ln == "{":
                in_block = True
            elif ln == "}":
                in_block = False
            elif in_block:
                out.append(ln)
        return out
    return [ln for ln in lines if ln not in ("{", "}")]


@app.route("/api/keywords/summary", methods=["GET"])
def keywords_summary():
    market = request.args.get("market") or None
    return jsonify(shopee_db.keyword_summary(DB_PATH, market=market))


@app.route("/api/keywords/list", methods=["GET"])
def keywords_list():
    market = request.args.get("market") or None
    status = request.args.get("status") or None
    search = request.args.get("search") or None
    try:
        cat_id = int(request.args["cat_id"]) if request.args.get("cat_id") not in (None, "") else None
    except (TypeError, ValueError):
        return _bad_request("'cat_id' phai la so nguyen")
    limit = int(request.args.get("limit") or 300)
    rows = shopee_db.list_keywords(DB_PATH, market=market, status=status,
                                   search=search, cat_id=cat_id, limit=limit)
    return jsonify({"keywords": rows})


@app.route("/api/keywords/import", methods=["POST"])
def keywords_import():
    """Nhap 1 LO tu khoa (cung market + cung cat_id/cat_name cho CA LO - dung quyet dinh
    "gắn cả lô khi người dùng ấn nút import"). CAC THONG SO SORT/FILTER/SOLD/HOA HONG KHONG
    NHAN O DAY - chung la cau hinh KHI CAO (worker gui kem moi /api/keywords/page_done).
    body: {market, keywords: [..] hoac text: "...", cat_id?, cat_name?}."""
    body = request.get_json(force=True, silent=True) or {}
    market = (body.get("market") or "").strip().lower()
    if market not in KEYWORD_MARKETS:
        return _bad_request(f"'market' phai la 1 trong: {', '.join(KEYWORD_MARKETS)}")
    keywords = body.get("keywords")
    if not isinstance(keywords, list):
        keywords = _parse_keyword_text(body.get("text"))
    if not isinstance(keywords, list) or not keywords:
        return _bad_request("thieu tu khoa: gui 'keywords' (list) hoac 'text' (nhieu dong)")
    try:
        cat_id = int(body["cat_id"]) if body.get("cat_id") not in (None, "") else None
    except (TypeError, ValueError):
        return _bad_request("'cat_id' phai la so nguyen")
    cat_name = (body.get("cat_name") or "").strip() or None
    if cat_id is not None and not cat_name:
        cat_name = shopee_categories.cat_name_for(market, cat_id)
    result = shopee_db.import_keywords(
        DB_PATH, market, keywords, cat_id=cat_id, cat_name=cat_name,
    )
    return jsonify({"ok": True, **result})


@app.route("/api/keywords/claim", methods=["POST"])
def keywords_claim():
    """Worker goi khi ranh: nhan 1 tu khoa pending (hoac con sot/in_progress lease het)
    DUNG market cua tab dang mo. Tra ve {'keyword': {...}} hoac keyword=null khi het viec."""
    body = request.get_json(force=True, silent=True) or {}
    device_key = (body.get("device_key") or "").strip()
    market = (body.get("market") or "").strip().lower()
    if not device_key:
        return _bad_request("thieu 'device_key'")
    if market not in KEYWORD_MARKETS:
        return _bad_request(f"'market' phai la 1 trong: {', '.join(KEYWORD_MARKETS)}")
    row = shopee_db.claim_keyword(DB_PATH, device_key, market)
    return jsonify({"keyword": row})


@app.route("/api/keywords/page_done", methods=["POST"])
def keywords_page_done():
    """Worker nop 1 trang (page_offset) item that da chup tu chinh trang affiliate. Server
    loc (theo CAU HINH CAO worker gui kem: sold_min/comm_money_min/filter_types - khong luu
    o import) + insert root pending moi (dedup toan DB) + cap nhat checkpoint/keyword."""
    body = request.get_json(force=True, silent=True) or {}
    keyword_id = body.get("keyword_id")
    device_key = (body.get("device_key") or "").strip()
    market = (body.get("market") or "").strip().lower()
    items = body.get("items")
    if keyword_id in (None, ""):
        return _bad_request("thieu 'keyword_id'")
    if not device_key:
        return _bad_request("thieu 'device_key'")
    if not market:
        return _bad_request("thieu 'market'")
    result = shopee_db.keyword_page_done(
        DB_PATH, keyword_id, device_key, market,
        body.get("page_offset"), body.get("page_limit"),
        body.get("total_count"), items if isinstance(items, list) else [],
        sold_min=body.get("sold_min"), comm_money_min=body.get("comm_money_min"),
        filter_types=body.get("filter_types"),
    )
    if not result.get("ok"):
        return jsonify(result), 409
    return jsonify(result)


@app.route("/api/keywords/<int:keyword_id>/fail", methods=["POST"])
def keywords_fail(keyword_id):
    body = request.get_json(force=True, silent=True) or {}
    reason = (body.get("reason") or "unknown_error")[:500]
    shopee_db.fail_keyword(DB_PATH, keyword_id, reason)
    return jsonify({"ok": True})


@app.route("/api/keywords/<int:keyword_id>/reset", methods=["POST"])
def keywords_reset_one(keyword_id):
    ok = shopee_db.reset_keyword(DB_PATH, keyword_id)
    if not ok:
        return _bad_request("khong tim thay keyword nay")
    return jsonify({"ok": True})


@app.route("/api/keywords/<int:keyword_id>", methods=["DELETE"])
def keywords_delete_one(keyword_id):
    ok = shopee_db.delete_keyword(DB_PATH, keyword_id)
    if not ok:
        return _bad_request("khong tim thay keyword nay")
    return jsonify({"ok": True})


@app.route("/api/keywords/bulk_reset", methods=["POST"])
def keywords_bulk_reset():
    """Dat lai hang loat tu khoa khop bo loc ve pending (bo loc rong = tat ca)."""
    body = request.get_json(force=True, silent=True) or {}
    n = shopee_db.reset_keywords(DB_PATH, market=body.get("market") or None,
                                 status=body.get("status") or None)
    return jsonify({"ok": True, "reset": n})


@app.route("/api/keywords/bulk_delete", methods=["POST"])
def keywords_bulk_delete():
    """Xoa hang loat tu khoa khop bo loc (KHONG dong cham toi root da bom)."""
    body = request.get_json(force=True, silent=True) or {}
    n = shopee_db.delete_keywords(DB_PATH, market=body.get("market") or None,
                                  status=body.get("status") or None)
    return jsonify({"ok": True, "deleted": n})


# ============================================================================
# Tab "Vận hành GPM" - dieu phoi worker cào qua GPM Login (Local API 9495).
# Server lam proxy (GPM khong CORS) + quan ly tien trinh node cdp_worker.mjs.
# ============================================================================
import requests as _requests  # noqa: E402

GPM_BASE = os.environ.get("GPM_BASE", "http://127.0.0.1:9495")
GPM_PORT_START = int(os.environ.get("GPM_PORT_START", "9601"))
_gpm_ports = {}     # profile_id -> cdp port da cap
_gpm_workers = {}   # profile_id -> {proc, name, market, port, log}
_gpm_lock = threading.RLock()  # RLock: _gpm_alloc_port() giu lock khi duoc goi tu trong handler cung lock


def _gpm_api(method, path, params=None, timeout=10):
    r = _requests.request(method, GPM_BASE + path, params=params, timeout=timeout)
    r.raise_for_status()
    try:
        return r.json()
    except ValueError:
        return {"success": False, "message": r.text[:200]}


def _gpm_tcp_up(port):
    if not port:
        return False
    try:
        with socket.create_connection(("127.0.0.1", int(port)), timeout=0.3):
            return True
    except OSError:
        return False


def _gpm_alloc_port(profile_id):
    with _gpm_lock:
        if profile_id in _gpm_ports:
            return _gpm_ports[profile_id]
        used = set(_gpm_ports.values())
        p = GPM_PORT_START
        tries = 0
        while (p in used or _gpm_tcp_up(p)) and tries < 200:
            p += 1
            tries += 1
        _gpm_ports[profile_id] = p
        return p


def _gpm_log_path(name):
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", str(name))
    log_dir = os.path.join(REPO_ROOT, "artifacts")
    os.makedirs(log_dir, exist_ok=True)
    return os.path.join(log_dir, f"gpm_worker_{safe}.log")


def _find_node():
    """Tim node.exe de spawn worker - khong chi tin PATH (server co the duoc khoi dong tu moi
    truong thieu node trong PATH, gay [WinError 2] 'Khong spawn duoc worker')."""
    cands = []
    try:
        w = shutil.which("node")
        if w:
            cands.append(w)
    except Exception:
        pass
    env_dirs = []
    for key in ("ProgramFiles", "ProgramFiles(x86)"):
        v = os.environ.get(key)
        if v:
            env_dirs.append(os.path.join(v, "nodejs", "node.exe"))
    la = os.environ.get("LOCALAPPDATA")
    if la:
        env_dirs.append(os.path.join(la, "Programs", "nodejs", "node.exe"))
    for p in env_dirs:
        if p and os.path.isfile(p) and p not in cands:
            cands.append(p)
    return cands[0] if cands else "node"


def _gpm_worker_status(profile_id):
    info = _gpm_workers.get(profile_id)
    if not info:
        return None
    proc = info.get("proc")
    running = proc is not None and proc.poll() is None
    return {
        "profile_id": profile_id,
        "name": info.get("name"),
        "market": info.get("market"),
        "mode": info.get("mode") or "root",
        "port": info.get("port"),
        "running": running,
        "exit_code": None if (proc is None or running) else proc.poll(),
        "started_at": info.get("started_at"),
    }


def _gpm_read_log_tail(name, n=40):
    path = _gpm_log_path(name)
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
        return "".join(lines[-n:])
    except OSError:
        return "(chua co log)"


def _gpm_kill_proc(proc):
    if not proc or proc.poll() is not None:
        return
    try:
        proc.terminate()
        try:
            proc.wait(timeout=4)
        except subprocess.TimeoutExpired:
            subprocess.run(
                ["taskkill", "/PID", str(proc.pid), "/T", "/F"],
                capture_output=True, timeout=10,
            )
    except Exception:
        pass


@app.route("/api/gpm/groups", methods=["GET"])
def gpm_groups():
    """Danh sach nhom (group) cua GPM - de UI chon nhom roi moi load profile cua nhom do."""
    try:
        data = _gpm_api("GET", "/api/v1/groups")
    except Exception as e:
        return jsonify({"ok": False, "error": f"Khong goi duoc GPM Local API ({GPM_BASE}): {e}"}), 502
    raw = data.get("data")
    if isinstance(raw, dict):
        items = raw.get("data") or []
    else:
        items = raw or []
    groups = [{"id": g.get("id"), "name": g.get("name") or g.get("id")} for g in items if g.get("id")]
    return jsonify({"ok": True, "groups": groups})


@app.route("/api/gpm/profiles", methods=["GET"])
def gpm_profiles():
    """Danh sach profile GPM + trang thai (port, worker, CDP). Loc theo ?group_id=<id> neu co."""
    group_id = (request.args.get("group_id") or "").strip()
    try:
        data = _gpm_api("GET", "/api/v1/profiles")
    except Exception as e:
        return jsonify({"ok": False, "error": f"Khong goi duoc GPM Local API ({GPM_BASE}): {e}"}), 502
    profiles = (data.get("data") or {}).get("data") or []
    if group_id:
        profiles = [p for p in profiles if str(p.get("group_id") or "") == str(group_id)]
    out = []
    for p in profiles:
        pid = p["id"]
        port = _gpm_ports.get(pid)
        st = _gpm_worker_status(pid)
        out.append({
            "id": pid,
            "name": p.get("name"),
            "group_id": p.get("group_id"),
            "market": "ph",  # mac dinh; sua trong UI khi chay
            "port": port,
            "cdp_up": _gpm_tcp_up(port) if port else False,
            "worker": st,
            "browser": (p.get("browser") or {}).get("name", "chrome"),
        })
    return jsonify({"ok": True, "gpm_base": GPM_BASE, "group_id": group_id or None, "profiles": out})


@app.route("/api/gpm/worker/start", methods=["POST"])
def gpm_worker_start():
    """Spawn 1 worker cho 1 profile GPM (tu start browser qua GPM khi chay).
    mode='root' (mac dinh): cdp_worker.mjs - cao tung root (offer/product/<item_id>).
    mode='keyword': cdp_keyword_worker.mjs - cao root AFF theo TU KHOA (worker claim tu
    khoa pending cua market, dieu khien chinh trang affiliate search + chup product/list)."""
    body = request.get_json(force=True, silent=True) or {}
    profile_id = (body.get("profile_id") or "").strip()
    name = (body.get("name") or profile_id).strip()
    market = (body.get("market") or "ph").strip()
    mode = (body.get("mode") or "root").strip()
    max_roots = int(body.get("max_roots") or 0)
    hidden = bool(body.get("hidden"))
    if not profile_id:
        return _bad_request("thieu 'profile_id'")
    if mode not in ("root", "keyword"):
        return _bad_request("'mode' chi nhan 'root' hoac 'keyword'")
    # Cau hinh "khi cào" (chi dung cho mode keyword): sort_type/filter_types (API param khi
    # search). Lượt bán & Hoa hồng KHONG gui o day - lay tu "Điều kiện lọc chung" (settings,
    # tab Worker GPM Login) phia server khi loc page_done (xem keyword_page_done).
    try:
        crawl_sort = int(body.get("sort_type") or 2)
        if crawl_sort not in (1, 2):
            crawl_sort = 2
        crawl_filter = int(body.get("filter_types") or 0)
    except (TypeError, ValueError):
        return _bad_request("'sort_type'/'filter_types' phai la so")
    worker_script = "cdp_worker.mjs" if mode == "root" else "cdp_keyword_worker.mjs"
    print(f"[gpm] start worker {name} ({profile_id}) mode={mode} crawl(sort={crawl_sort}, filter={crawl_filter})...", flush=True)
    with _gpm_lock:
        st = _gpm_worker_status(profile_id)
        if st and st["running"]:
            return jsonify({"ok": False, "error": f"Worker '{st['name']}' dang chay roi (pid da co)."}), 409
        port = _gpm_alloc_port(profile_id)
        log_path = _gpm_log_path(name + ("_keyword" if mode == "keyword" else ""))
        node_exe = _find_node()
        cmd = [
            node_exe,
            os.path.join(SCRIPTS_DIR, worker_script),
            "--gpm-profile", profile_id,
            "--port", str(port),
            "--device-key", name,
            "--market", market,
            "--log", log_path,
        ]
        if mode == "keyword":
            cmd += ["--sort-type", str(crawl_sort), "--filter-types", str(crawl_filter)]
        if max_roots > 0:
            limit_arg = "--max-keywords" if mode == "keyword" else "--max-roots"
            cmd += [limit_arg, str(max_roots)]
        if hidden:
            cmd += ["--hidden", "1"]
    # spawn NGOAI lock de khong chan cac request khac
    print(f"[gpm] spawn node: {' '.join(cmd)}", flush=True)
    logf = open(log_path, "a", encoding="utf-8")
    try:
        proc = subprocess.Popen(cmd, cwd=REPO_ROOT, stdout=logf, stderr=subprocess.STDOUT, text=True)
    except Exception as e:
        logf.close()
        return jsonify({"ok": False, "error": f"Khong spawn duoc worker: {e}"}), 500
    _gpm_workers[profile_id] = {
        "proc": proc, "name": name, "market": market, "mode": mode,
        "port": port, "log": log_path,
        "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    print(f"[gpm] da spawn worker {name} (profile {profile_id}) mode={mode} pid={proc.pid} port={port}")
    return jsonify({"ok": True, "worker": _gpm_worker_status(profile_id)})


@app.route("/api/gpm/worker/stop", methods=["POST"])
def gpm_worker_stop():
    body = request.get_json(force=True, silent=True) or {}
    profile_id = (body.get("profile_id") or "").strip()
    stop_browser = bool(body.get("stop_browser", True))
    if not profile_id:
        return _bad_request("thieu 'profile_id'")
    info = _gpm_workers.get(profile_id)
    if info:
        _gpm_kill_proc(info["proc"])
        info["proc"] = None
    msg = "Da dung worker."
    if stop_browser:
        try:
            _gpm_api("GET", f"/api/v1/profiles/stop/{profile_id}")
            msg += " + da dong browser GPM."
        except Exception as e:
            msg += f" (dong browser loi: {e})"
    return jsonify({"ok": True, "message": msg})


@app.route("/api/gpm/worker/log", methods=["GET"])
def gpm_worker_log():
    name = (request.args.get("name") or "").strip()
    if not name:
        return _bad_request("thieu 'name'")
    # Log cua worker keyword duoc ghi vao file khac (co duoi _keyword) nen tim THEO worker
    # dang chay truoc; fallback lai duong dan mac dinh theo ten (mode root / worker cu).
    info = next((v for v in _gpm_workers.values() if v.get("name") == name), None)
    log_path = info.get("log") if info and info.get("log") else _gpm_log_path(name)
    try:
        with open(log_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
        text = "".join(lines[-60:])
    except OSError:
        text = "(chua co log)"
    return jsonify({"ok": True, "log": text})


_GPM_HOME_URL = {
    "ph": "https://shopee.ph/",
    "th": "https://shopee.co.th/",
    "my": "https://shopee.com.my/",
    "vn": "https://shopee.vn/",
    "sg": "https://shopee.sg/",
    "id": "https://shopee.co.id/",
}


def _gpm_ensure_browser(profile_id):
    """Bao dam browser GPM cua profile dang chay + tra ve CDP port that su (tu GPM tra ve khi
    start - khong tu chon port de tranh lech voi instance dang chay san). Neu GPM bao
    ProfileInUse (dang chay tu noi khac nhung khong ro port) thi stop roi start lai 1 lan."""
    with _gpm_lock:
        known = _gpm_ports.get(profile_id)
        if known and _gpm_tcp_up(known):
            return known
    for attempt in range(2):
        try:
            r = _gpm_api("GET", f"/api/v1/profiles/start/{profile_id}")
        except Exception:
            return None
        if r.get("success"):
            data = r.get("data") or {}
            p = data.get("remote_debugging_port")
            if p:
                with _gpm_lock:
                    _gpm_ports[profile_id] = p
                return p
            return None
        if "InUse" in (r.get("message") or "") and attempt == 0:
            try:
                _gpm_api("GET", f"/api/v1/profiles/stop/{profile_id}")
            except Exception:
                pass
            time.sleep(2.5)
            continue
        return None
    return None


@app.route("/api/gpm/browser/open", methods=["POST"])
def gpm_browser_open():
    """Mo browser GPM cua profile (neu chua chay) va mo 1 tab toi 'url' - dung cho nut
    'Home-Shopee' (mo trang chu shopee.<market> theo market dang chon cua profile)."""
    body = request.get_json(force=True, silent=True) or {}
    profile_id = (body.get("profile_id") or "").strip()
    url = (body.get("url") or "").strip()
    market = (body.get("market") or "").strip()
    if not profile_id:
        return _bad_request("thieu 'profile_id'")
    if not url and market in _GPM_HOME_URL:
        url = _GPM_HOME_URL[market]
    if not url or not (url.startswith("https://") or url.startswith("http://")):
        return _bad_request("thieu 'url' hop le")
    port = _gpm_ensure_browser(profile_id)
    if not port:
        return jsonify({"ok": False, "error": "GPM khong start duoc browser (kiem tra GPM app / profile dang mo)."}), 502
    up = False
    for _ in range(90):  # cho toi 45s browser bind CDP
        if _gpm_tcp_up(port):
            up = True
            break
        time.sleep(0.5)
    if not up:
        return jsonify({"ok": False, "error": f"Browser GPM start nhung CDP port {port} khong len."}), 502
    # mo tab moi bang CDP HTTP endpoint /json/new?<url> - CHU Y: url phai nam THANG trong query
    # (khong phai param ten 'url' - Chrome bo qua neu dung dang '?url=...' va chi mo about:blank)
    created = False
    new_tab = f"http://127.0.0.1:{port}/json/new?{url}"
    try:
        r = _requests.put(new_tab, timeout=6)
        if r.status_code in (200, 201):
            created = True
    except Exception:
        pass
    if not created:
        try:
            r = _requests.get(new_tab, timeout=6)
            if r.status_code in (200, 201):
                created = True
        except Exception:
            pass
    if not created:
        return jsonify({"ok": False, "error": f"Mo tab that bai tren port {port}."}), 502
    return jsonify({"ok": True, "url": url, "port": port, "message": f"Da mo {url}"})


def _ensure_port_free(host, port):
    """Bind THAT (roi dong ngay) truoc khi giao cho Werkzeug - phat hien SOM va bao loi RO
    RANG neu port da co server khac dang chay, thay vi de Werkzeug tu bind. Ly do: da xac
    nhan THAT tren may nay Windows cho phep 2 tien trinh CUNG bind duoc 1 port TCP ma KHONG
    bao loi "address already in use" (hanh vi bind mac dinh cua Werkzeug tren Windows) - khi
    do request tu trinh duyet bi he dieu hanh dinh tuyen NGAU NHIEN vao 1 trong 2 tien trinh.
    Neu tien trinh con lai "chet"/ket (vd dang cho 1 giao dich SQLite khong bao gio commit),
    request roi vao no se treo VINH VIEN - dung trieu chung nguoi dung bao cao: dashboard im
    lang khong phan hoi sau 1 luc, phai F5 lai. SO_EXCLUSIVEADDRUSE la co RIENG Windows ep he
    dieu hanh tu choi thang lan bind thu 2, bien loi ngam thanh loi ro rang ngay luc khoi
    dong thay vi lam hong ngau nhien luc dang dung (start_affiliate_scraper.bat da tung co
    kiem tra tuong tu qua netstat, nhung khong bao ve duoc neu server duoc khoi dong theo
    cach khac ngoai file .bat do)."""
    probe = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        if hasattr(socket, "SO_EXCLUSIVEADDRUSE"):
            probe.setsockopt(socket.SOL_SOCKET, socket.SO_EXCLUSIVEADDRUSE, 1)
        probe.bind((host, port))
    except OSError:
        print(
            f"[affiliate_scrape_server] LOI: port {port} DA CO server khac dang chay san "
            f"(vd mo .bat 2 lan, hoac 1 cua so terminal khac chua dong). Dung 2 tien trinh "
            f"cung chiem 1 port se gay loi 'dashboard khong phan hoi ngau nhien'. Dong server "
            f"cu (hoac dung cua so dang chay san) truoc khi mo cai moi."
        )
        sys.exit(1)
    finally:
        probe.close()


def main():
    global DB_PATH
    ap = argparse.ArgumentParser()
    ap.add_argument("--port", type=int, default=8877)
    ap.add_argument("--db-path", default=shopee_db.DB_PATH_DEFAULT)
    args = ap.parse_args()
    DB_PATH = args.db_path
    _ensure_port_free("127.0.0.1", args.port)
    shopee_db.init_db(DB_PATH)  # dam bao bang/cot ton tai truoc khi nhan request dau tien
    print(f"[affiliate_scrape_server] DB: {DB_PATH} | http://127.0.0.1:{args.port}")
    # threaded=True QUAN TRONG: mac dinh Werkzeug dev server xu ly TUAN TU tung request 1
    # (single-threaded) - voi so luong tab Tampermonkey (worker) chay song song + dashboard
    # tu poll 4 API moi 5s, request nao cung phai xep hang cho request truoc xong. Nguoi
    # dung bao cao trieu chung "bam nut tren dashboard khong phan hoi, giong mat mang" - dung
    # la hien tuong request bi ket trong hang doi nay, KHONG phai loi mang that. An toan bat
    # thread vi tang du lieu (shopee_db.py) da thiet ke san cho ghi song song: moi ham tu mo
    # 1 connection SQLite RIENG (_connect(), khong dung chung giua cac request/thread) + WAL
    # mode + BEGIN IMMEDIATE cho cac giao dich ghi quan trong (xem init_db()/try_assign_verified()).
    app.run(host="127.0.0.1", port=args.port, debug=False, threaded=True)


if __name__ == "__main__":
    main()
